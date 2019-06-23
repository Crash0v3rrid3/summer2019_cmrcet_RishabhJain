import click
import openpyxl
import os

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject.settings')
import django
django.setup()

from onlineapp.models import *

class ExcelWorkBook(click.ParamType): # custom type
    name = 'excel-work-book'

    def convert(self, value, param, ctx):
        wbName = value

        try:
            wb = openpyxl.load_workbook(wbName) # open it
        except FileNotFoundError:
            self.fail(f'Could not open file {self.wbName}') # if doesn't exist fail

        return wb


@click.group()
@click.pass_context
def init(ctx):
    pass

@init.command(name='createdb', help='command that creates a database and the appropriate tables')
@click.pass_context
def createdb(ctx):
    pass


@init.command(name='dropdb', help='Drop a database')
@click.argument('databasename', required=True, default='mrnd')
@click.pass_context
def dropdb(ctx, databasename):
    ctx.obj["conn"].query(f'DROP DATABASE IF EXISTS {databasename};')


@init.command(name='importdata', help='import data from an excel sheet')
@click.argument('source_students_colleges', type = ExcelWorkBook())
@click.argument('marks', type = ExcelWorkBook())
@click.pass_context
def importdata(ctx, source_students_colleges, marks):
    source_students_colleges.active = source_students_colleges.sheetnames.index('Colleges')
    sourceSheet = source_students_colleges.active

    for row in range(2, sourceSheet.max_row + 1):
        college = College(name = sourceSheet.cell(row=row,column=1).value,
                          location=sourceSheet.cell(row=row, column=3).value,
                          acronym=str(sourceSheet.cell(row=row, column=2).value).lower(),
                          contact=sourceSheet.cell(row=row, column=4).value)
        try:
            college.save()
        except:
            pass


    source_students_colleges.active = source_students_colleges.sheetnames.index('Current')
    sourceSheet = source_students_colleges.active

    for row in range(2, sourceSheet.max_row + 1):
        try:
            student = Student(name=sourceSheet.cell(row=row, column=1).value,
                              college=College.objects.get(acronym=str(sourceSheet.cell(row=row, column=2).value).lower()),
                              email=sourceSheet.cell(row=row, column=3).value,
                              db_folder=str(sourceSheet.cell(row=row, column=4).value).lower())

            student.save()
        except:
            pass

    source_students_colleges.active = source_students_colleges.sheetnames.index('Deletions')
    sourceSheet = source_students_colleges.active

    for row in range(2, sourceSheet.max_row + 1):
        try:
            student = Student(name=sourceSheet.cell(row=row, column=1).value,
                              college=College.objects.get(
                                  acronym=str(sourceSheet.cell(row=row, column=2).value).lower()),
                              email=sourceSheet.cell(row=row, column=3).value,
                              db_folder=str(sourceSheet.cell(row=row, column=4).value).lower(),
                              dropped_out=True)

            student.save()
        except:
            pass


    sourceSheet = marks.active

    for row in range(2, sourceSheet.max_row + 1):
        try:
            mark = MockTest1(student=Student.objects.get(db_folder=str(sourceSheet.cell(row=row, column=1).value).split('_')[2].lower()),
                             problem1=sourceSheet.cell(row=row, column=2).value,
                             problem2=sourceSheet.cell(row=row, column=3).value,
                             problem3=sourceSheet.cell(row=row, column=4).value,
                             problem4=sourceSheet.cell(row=row, column=5).value,
                             total=sourceSheet.cell(row=row, column=6).value)
            mark.save()
        except:
            pass


@init.command(name = 'collegestats', help = 'print college statistics')
@click.pass_context
def collegestats(ctx):
    conn = ctx.obj['conn']
    conn.select_db('mrnd')
    conn.query('SELECT S.college, MIN(M.total), MAX(M.total), AVG(M.total), COUNT(M.total) FROM marks M, students S WHERE M.dbname = S.dbname GROUP BY S.college ORDER BY S.college')

    result = conn.use_result()

    row = result.fetch_row()
    print('COLLEGE\tMIN\tMAX\tAVG\tTOTAL STUDENTS')
    while(row != ()):
        print('\t'.join(list(map(lambda x: x.decode('utf-8'),row[0]))))
        row = result.fetch_row()

if __name__ == '__main__':
    init()