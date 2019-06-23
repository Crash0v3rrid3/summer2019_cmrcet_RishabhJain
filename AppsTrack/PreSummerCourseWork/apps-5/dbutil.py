import click
from MySQLdb import _mysql
import openpyxl


class ExcelWorkBook(click.ParamType): # custom type
    name = 'excel-work-book'

    def convert(self, value, param, ctx):
        wbName = value

        try:
            wb = openpyxl.load_workbook(wbName) # open it
        except FileNotFoundError:
            self.fail(f'Could not open file {self.wbName}') # if doesn't exist fail

        return wb


@click.group()
@click.option('-h', '--host', default='127.0.0.1', help='mysql host')
@click.option('-u', '--username', default='root', help='mysql username')
@click.option('-P', '--password',  prompt=True, hide_input=True, confirmation_prompt=True, help='Password for db user')
@click.pass_context
def init(ctx, host, username, password):
    ctx.obj = {'conn': _mysql.connect(host, username, password)}

@init.command(name='createdb', help='command that creates a database and the appropriate tables')
@click.option('-dn', '--db-name', default='mrnd', help='specify database name', required=True)
@click.option('-t', '--table', help='specify tables to create in format "tablename(column1 datatype1, column2 datatype2...)"', nargs = 2, required=True)
@click.option('-f', '--foreign-key', help='Specify forien key attributes like "table1.key,table2.key"', required=True)
@click.pass_context
def createdb(ctx, db_name, table, foreign_key):
    ctx.obj["conn"].query(f'DROP DATABASE IF EXISTS {db_name}')
    ctx.obj["conn"].query(f'CREATE DATABASE {db_name};')
    #database created

    ctx.obj["conn"].select_db(db_name)
    #database selected

    for singleTable in table:
        ctx.obj["conn"].query(f'CREATE TABLE {singleTable}')
        #ables created

    tables = tuple(map(lambda x: x.split('.'), foreign_key.split(',')))
    #[[table1, key1], [table2, key2]]
    query = f'ALTER TABLE {db_name}.{tables[0][0]} ADD FOREIGN KEY ({tables[0][1]}) REFERENCES {db_name}.{tables[1][0]}({tables[1][1]})'
    ctx.obj["conn"].query(query)
    #foreign keys added


@init.command(name='dropdb', help='Drop a database')
@click.argument('databasename', required=True, default='mrnd')
@click.pass_context
def dropdb(ctx, databasename):
    ctx.obj["conn"].query(f'DROP DATABASE IF EXISTS {databasename};')

def getFields(source):
    sourceSheet = source.active
    fields = (sourceSheet.cell(row = 1, column = column) for column in range(1, sourceSheet.max_column + 1))

    return fields

@init.command(name='importdata', help='import data from an excel sheet')
@click.argument('source', type = ExcelWorkBook())
@click.argument('databasename', required=True, default='mrnd')
@click.option('-t', '--table-name', help='table name to which to insert data')
@click.option('-s', '--sheet-name', help='Sheet name in workbook')
@click.pass_context
def importdata(ctx, table_name, source, sheet_name, databasename):
    #fields = getFields(source)
    ctx.obj["conn"].select_db(databasename)

    source.active = source.sheetnames.index(sheet_name)
    sourceSheet = source.active


    for row in range(2, sourceSheet.max_row):
        if table_name != 'marks':
            fields = str(('name', 'college', 'email', 'dbname')).replace("'", '')
            values = [str(sourceSheet.cell(row = row, column = column).value) for column in range(1, sourceSheet.max_column + 1)]
            values[-1] = values[-1].lower()
        else:
            fields = str(('foldername', 'transform', 'from_custom_base26', 'get_pig_latin', 'top_chars', 'total', 'dbname')).replace('\'', '')
            values = [str(sourceSheet.cell(row=row, column=column).value) for column in
                      range(1, sourceSheet.max_column + 1)]
            values.append(values[0].split('_')[2])

        values = str(tuple(values))
        query = f'INSERT INTO {table_name}{fields} VALUES{values};'

        try:
            ctx.obj["conn"].query(query)
        except:
            pass

@init.command(name = 'collegestats', help = 'print college statistics')
@click.pass_context
def collegestats(ctx):
    conn = ctx.obj['conn']
    conn.select_db('mrnd')
    conn.query('SELECT S.college, MIN(M.total), MAX(M.total), AVG(M.total), COUNT(M.total) FROM marks M, students S WHERE M.dbname = S.dbname GROUP BY S.college ORDER BY S.college')

    result = conn.use_result()

    row = result.fetch_row()
    print('COLLEGE\tMIN\tMAX\tAVG\tTOTAL STUDENTS')
    while(row != ()):
        print('\t'.join(list(map(lambda x: x.decode('utf-8'),row[0]))))
        row = result.fetch_row()

if __name__ == '__main__':
    init()