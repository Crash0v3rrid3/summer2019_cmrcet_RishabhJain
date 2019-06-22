import openpyxl
import django
import os
import datetime

# setup django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'appday.settings')
django.setup()
from iplclone.models import *

# open sheet
matches = openpyxl.load_workbook('matches.xlsx')
deliveries = openpyxl.load_workbook('deliveries.xlsx')

matches_sheet = matches.active
deliveries_sheet = deliveries.active


def extractDate(dateString):

    if '-' in dateString:
        year, month, day = map(int, dateString.split('-'))
    else:
        day, month, year = map(int, dateString.split('/'))
        year += 2000
    dateObj = datetime.datetime(year, month, day)

    return dateObj


def findMatch(matchTuple):
    match = matchTuple
    try:
        team1 = Team.objects.get(name=match[3])
        team2 = Team.objects.get(name=match[4])
        toss_winner = Team.objects.get(name=match[5])


        return Match.objects.get(
            season=Season.objects.get(year=int(match[0])),
            team1=team1,
            team2=team2,
            date=extractDate(match[2]),
        )

    except Exception as err:
        pass

def findRowInMatches(id):
    global matches_sheet

    for row in range(2, matches_sheet.max_row + 1):
        if id == matches_sheet.cell(row=row, column=1).value:
            return tuple(map(lambda x: x.value, tuple(matches_sheet[row])[1:]))

    return None


rows = [[y.value for y in x] for x in matches_sheet.rows]
rows = rows[1: ]
matches = []

for row in rows:
    rowTuple = row

    try:
        match = Match(
            team1=Team.objects.get(name=rowTuple[4]),
            team2=Team.objects.get(name=rowTuple[5]),
            season=Season.objects.get(year=int(rowTuple[1])),
            winner=(rowTuple[4] == rowTuple[10]),
            venue=Venue.objects.get(stadium__name=rowTuple[14]),
            date=extractDate(rowTuple[3]),
            toss_winner=(rowTuple[4] == rowTuple[6]),
            toss_decision=(rowTuple[7] == 'field'),
            result=Result.objects.get(result=rowTuple[8]),
            win_by_runs=int(rowTuple[11]),
            win_by_wickets=int(rowTuple[12]),
            dl_applied=rowTuple[9],
            player_of_match=(Player.objects.get(name=rowTuple[13]) if rowTuple[13] else None),
            umpire1=(Umpire.objects.get(name=rowTuple[15]) if rowTuple[15] else None),
            umpire2=(Umpire.objects.get(name=rowTuple[16]) if rowTuple[16] else None),
            umpire3=(Umpire.objects.get(name=rowTuple[17]) if rowTuple[17] else None),
            tmp_match_id=rowTuple[0]
        )

        matches.append(match)
    except Exception as err:
        print(err, rowTuple)
        pass


Match.objects.bulk_create(matches)