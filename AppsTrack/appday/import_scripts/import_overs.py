import openpyxl
import django
import os
import datetime

# setup django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'appday.settings')
django.setup()
from iplclone.models import *

# open sheet
matches = openpyxl.load_workbook('matches.xlsx')
deliveries = openpyxl.load_workbook('deliveries.xlsx')

matches_sheet = matches.active
deliveries_sheet = deliveries.active


def extractDate(dateString):

    if '-' in dateString:
        year, month, day = map(int, dateString.split('-'))
    else:
        day, month, year = map(int, dateString.split('/'))
        year += 2000
    dateObj = datetime.datetime(year, month, day)

    return dateObj


def findMatch(matchTuple):
    match = matchTuple
    try:
        team1 = Team.objects.get(name=match[3])
        team2 = Team.objects.get(name=match[4])
        toss_winner = Team.objects.get(name=match[5])


        return Match.objects.get(
            season=Season.objects.get(year=int(match[0])),
            team1=team1,
            team2=team2,
            date=extractDate(match[2]),
        )

    except Exception as err:
        pass

def findRowInMatches(id):
    global matches_sheet

    for row in range(2, matches_sheet.max_row + 1):
        if id == matches_sheet.cell(row=row, column=1).value:
            return tuple(map(lambda x: x.value, tuple(matches_sheet[row])[1:]))

    return None

id = None
inings = set()
rows = [[x[0].value, x[1].value, x[2].value, x[4].value, x[6].value, x[7].value, x[8].value, x[9].value] for x in deliveries_sheet.rows]
rows = rows[1: ]
overs = {}
matches = {}

for row in rows:
    rowTuple = row
    if id != rowTuple[0]:
        id = rowTuple[0]
        if matches.get(id) is None:
            matches[id] = Match.objects.get(tmp_match_id=id)
        match = matches[id]#findMatch(findRowInMatches(id))

    if match:
        try:
            ining = Ining.objects.get(
                match=match,
                inings=(rowTuple[1]),
                batting_team=Team.objects.get(name=rowTuple[2]),
            )

            if overs.get((rowTuple[0], rowTuple[1], rowTuple[3])) is None:
                overs[(match.id, rowTuple[1], rowTuple[3])] = Over(
                    ining=ining,
                    is_super_over=rowTuple[7],
                )

        except Exception as err:
            print(err, match.id, rowTuple)
            pass

Over.objects.bulk_create(overs.values())