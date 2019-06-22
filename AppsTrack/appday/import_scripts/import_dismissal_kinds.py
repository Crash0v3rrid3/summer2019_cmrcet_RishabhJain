import openpyxl
import django
import os
import datetime

# setup django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'appday.settings')
django.setup()
from iplclone.models import *

# open sheet
matches = openpyxl.load_workbook('matches.xlsx')
deliveries = openpyxl.load_workbook('deliveries.xlsx')

matches_sheet = matches.active
deliveries_sheet = deliveries.active


def extractDate(dateString):

    if '-' in dateString:
        year, month, day = map(int, dateString.split('-'))
    else:
        day, month, year = map(int, dateString.split('/'))
        year += 2000
    dateObj = datetime.datetime(year, month, day)

    return dateObj


def findMatch(matchTuple):
    match = matchTuple
    try:
        team1 = Team.objects.get(name=match[3])
        team2 = Team.objects.get(name=match[4])
        toss_winner = Team.objects.get(name=match[5])


        return Match.objects.get(
            season=Season.objects.get(year=int(match[0])),
            team1=team1,
            team2=team2,
            date=extractDate(match[2]),
        )

    except Exception as err:
        pass

def findRowInMatches(id):
    global matches_sheet

    for row in range(2, matches_sheet.max_row + 1):
        if id == matches_sheet.cell(row=row, column=1).value:
            return tuple(map(lambda x: x.value, tuple(matches_sheet[row])[1:]))

    return None

id = None
inings = set()
rows = [x[19].value for x in deliveries_sheet.rows]
rows = rows[1: ]
dismissal_kinds = set()

for dismissal_kind in rows:
    if dismissal_kind:
        dismissal_kinds.add(dismissal_kind)


DismissalKind.objects.bulk_create(list(map(lambda x: DismissalKind(name=x), dismissal_kinds)))
